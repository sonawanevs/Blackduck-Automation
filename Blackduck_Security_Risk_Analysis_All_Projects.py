import requests
import json
import re
import os
import datetime
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter

# API Key
bd_api_key = 'token XXXXXXXXXXXXXXXXXXXXXXXX'

# Connect to Blackduck via API key and get the bearer token
bd_url = 'https://demo.yourblackducksite.com/api/tokens/authenticate'
headers = {'Authorization': bd_api_key}
bdresponse = requests.post(bd_url, headers=headers)

if bdresponse.status_code == 200:
    blackduck_response = bdresponse.text
    print(blackduck_response)
    # print (blackduck_response)
    print("\nBlackduck is up and running! You have been authenticated via your API key.\n")

    # Step 1: Capture Bearer token for further requests
    bearertk = re.findall('\{"bearerToken":"(.*?)".*', blackduck_response)

    # Step 2: Form Authorization header with bearer token
    bd_bearer_token = 'Bearer ' + str(''.join(bearertk))
    headers_2 = {'Authorization': bd_bearer_token}

    # Step 3: Make API call with bearer token and fetch project ID and version ID
    project_url = 'https://demo.yourblackducksite.com/api/projects/?limit=5000'
    project_response = requests.get(project_url, headers=headers_2)
    project_names = ()
    project_names = (re.findall(r'"name":"(.*?)".*?"href":"(.*?)"', str(project_response.text), re.M))

    print("The following application results are available:\n")

    counter = 1
    while counter <= (len(project_names)):
        print(str(counter) + ': ' + (project_names[counter-1][0]))
        counter = counter + 1

    project_index_number = 0
    while project_index_number <= (len(project_names)):
        print("Starting analysis for the application : " + str(project_names[int(project_index_number)]))
        print("\n")
        # get project_id_number
        project_name_id = (re.findall(r'.*/(.*)', str(project_names[int(project_index_number)][1]), re.M))

        # Step 3: Make API call to get Project versions
        version_url = str(project_names[int(project_index_number)][1]) + '/versions'
        version_response = requests.get(version_url, headers=headers_2)
        version_names = ()
        version_names = (
            re.findall(r'.*"createdAt":"(.*?)".*?"versionName":"(.*?)".*?"href":"(.*?)"', str(version_response.text), re.M))
        print("Selected Application Code Release Version : " + str(version_names[0][1]) + ", Scan Date: " + str(
            version_names[0][0]))
        print("\n")
        project_version_id = re.findall(r'.*/(.*)', version_names[0][2], re.M)

        bd_project_id = str(project_name_id[0])
        bd_project_version_id = str(project_version_id[0])

        bom_url = 'https://demo.yourblackducksite.com/api/v1/releases/' + bd_project_version_id + '/component-bom-entries?limit=5000'

        # Step 3: Make API call with bearer token
        bom_url_response = requests.get(bom_url, headers=headers_2)
        fh = open("bom.txt", "w")
        fh.write(bom_url_response.text)
        fh.close()

        fh = open("bom.txt")
        bom_content = fh.read()
        fh.close()

        # bom_components=[]
        bom_components = ()

        # Multiple Occurances..
        bom_components = (re.findall(
            r'"(projectId)":"(.*?)".*?"(projectName)":"(.*?)".*?"(releaseId)":"(.*?)".*?"(releaseVersion)":"(.*?)".*?"(usages)":\["(.*?)"\].*?"(matchTypes)":\["(.*?)"\]',
            bom_content, re.M))

        # OPERATIONAL RISK PROFILE
        bom_risk_operational_profile = ()
        bom_risk_operational_profile = (re.findall(r'"(projectName)":"(.*?)".*?"(OPERATIONAL)":{(.*?)}', bom_content, re.M))
        # print(bom_risk_operational_profile)

        # VULNERABILITY RISK PROFILE
        bom_risk_vuln_profile = ()
        bom_risk_vuln_profile = (re.findall(r'"(projectName)":"(.*?)".*?"(VULNERABILITY)":{(.*?)}', bom_content, re.M))
        # print(bom_risk_vuln_profile[0])

        # LICENSE RISK PROFILE
        bom_risk_license_profile = ()
        bom_risk_license_profile = (
            re.findall(r'"(projectName)":"(.*?)".*?"licenses":.*?"(name)":"(.*?)".*?"(LICENSE)":{(.*?)}', bom_content,
                       re.M))

        # Create Component ID Text File
        counterr_1 = 0
        f = open("component_ID.txt", "w")
        while counterr_1 < (len(bom_components)):
            if (counterr_1 < ((len(bom_components) - 1))):
                f.write(str(bom_components[counterr_1][1]))
                f.write("\n")
            else:
                f.write(str(bom_components[counterr_1][1]))
            counterr_1 += 1
        f.close()

        # Create Component version ID Text File
        counterr_2 = 0
        f = open("component_version_ID.txt", "w")
        while counterr_2 < (len(bom_components)):
            if (counterr_2 < ((len(bom_components) - 1))):
                f.write(str(bom_components[counterr_2][5]))
                f.write("\n")
            else:
                f.write(str(bom_components[counterr_2][5]))
            counterr_2 += 1
        f.close()

        # Create Component Name Text File
        counterr_3 = 0
        f = open("component_name.txt", "w")
        while counterr_3 < (len(bom_components)):
            if (counterr_3 < ((len(bom_components) - 1))):
                f.write(str(bom_components[counterr_3][3]))
                f.write("\n")
            else:
                f.write(str(bom_components[counterr_3][3]))
            counterr_3 += 1
        f.close()

        # ----------------------------------------------------------------------------------------------------------------
        #### START OF CAPTURING CVSS SCORE
        # Step 3: Make API call with bearer token
        vul_bom_url = 'https://demo.yourblackducksite.com/api/projects/' + bd_project_id + '/versions/' + bd_project_version_id + '/vulnerable-bom-components?limit=10000'
        vul_bom_url_response = requests.get(vul_bom_url, headers=headers_2)
        fh = open("vulbom.txt", "w")
        fh.write(vul_bom_url_response.text)
        fh.close()

        fh = open("vulbom.txt")
        vul_bom_content = fh.read()
        fh.close()

        # bom_components=[]
        vul_bom_components = []
        vul_bom_components = re.findall(
            r'{.*?"componentName":"(.*?)".*?"componentVersionName":"(.*?)".*?"baseScore":(.*?),', vul_bom_content, re.M)

        name_version = []
        counter = 0
        while counter < (len(vul_bom_components)):
            name_version_str = vul_bom_components[counter][0] + '-' + vul_bom_components[counter][1]
            name_version.append(name_version_str)
            counter = counter + 1

        # intilize a null list to capture unique components from the list
        unique_list = []
        unique_comp_list = []
        counter = 0
        # traverse for all elements
        for x in (name_version):
            if x not in unique_comp_list:
                unique_comp_list.append(x)

        # traverse through vul comp and capture all CVSS scores
        cvss_high_score = []
        cvss = []
        cvss_comp_name = []
        a = 0
        y = 0
        for x in unique_comp_list:
            a = 0
            v = '0.0'
            cvss.append(x)
            cvss.append(v)
            while a < len(name_version):
                if x == name_version[a]:
                    if cvss[1] < (vul_bom_components[a][2]):
                        cvss[1] = vul_bom_components[a][2]
                a = a + 1
            cvss_comp_name.append(x)
            cvss_high_score.append(cvss)
            cvss = []


        # --------------------------------- End of Phase 1 ------------------------------------------------------

        # --------------------------------- PHASE 2-----------------------------------------

        # -------------------------------------------- FUNCTION DEFENIIONS ----------------------------------------
        # Functions:
        # 1. Parse Component Risk
        def compoent_risk_profile(risk_profile_url_response):
            risk_profile_details = []

            # 3.1.1 - Release Date
            comp_release_date = re.findall('.*{"releasedOn":"(.*?)T.*".*', risk_profile_url_response)
            if (len(comp_release_date) > 0):
                risk_profile_details.append(comp_release_date.pop(0))
            else:
                risk_profile_details.append("-")

                # 3.1.2 - New Revisions
            comp_new_release_counts = re.findall('.*"newerReleasesCount":(.*?)}.*', risk_profile_url_response)
            if (len(comp_new_release_counts) > 0):
                risk_profile_details.append(comp_new_release_counts.pop(0))
            else:
                risk_profile_details.append("-")

                # 3.1.3 - Commit Trend
            comp_commit_trend = re.findall('.*"trending":"(.*?)"}.*', risk_profile_url_response)
            if (len(comp_commit_trend) > 0):
                risk_profile_details.append(comp_commit_trend.pop(0))
            else:
                risk_profile_details.append("-")

                # 3.1.4 - Last Commit Date
            comp_last_commit_date = re.findall('.*"lastCommitDate":"(.*?)T.*",.*}.*', risk_profile_url_response)
            if (len(comp_last_commit_date) > 0):
                risk_profile_details.append(comp_last_commit_date.pop(0))
            else:
                risk_profile_details.append("-")

                # 3.1.5 - Vulnerabilities (High, Medium and Low)
            comp_high_vulnerabilities = re.findall('.*"HIGH","count":(.*?)},.*}.*', risk_profile_url_response)
            if (len(comp_high_vulnerabilities) > 0):
                risk_profile_details.append(comp_high_vulnerabilities.pop(0))
            else:
                risk_profile_details.append("-")

            comp_medium_vulnerabilities = re.findall('.*"MEDIUM","count":(.*?)},.*}.*', risk_profile_url_response)
            if (len(comp_medium_vulnerabilities) > 0):
                risk_profile_details.append(comp_medium_vulnerabilities.pop(0))
            else:
                risk_profile_details.append("-")

            comp_low_vulnerabilities = re.findall('.*"LOW","count":(.*?)},.*}.*', risk_profile_url_response)
            if (len(comp_low_vulnerabilities) > 0):
                risk_profile_details.append(comp_low_vulnerabilities.pop(0))
            else:
                risk_profile_details.append("-")

            return risk_profile_details;


        # 2. Parse remediation version
        def remedy_version(remedy_version_response):
            remedy_version_details = []
            # print (remedy_version_response)
            remedy_version_number = re.findall(
                '{"fixesPreviousVulnerabilities":{"componentVersion":".*?","name":"(.*?)".*}', remedy_version_response)

            if (len(remedy_version_number) > 0):
                # print (remedy_version_number)
                remedy_version_details.append(remedy_version_number.pop(0))
            else:
                remedy_version_number = re.findall('.*noVulnerabilities":{"componentVersion":".*?","name":"(.*?)".*}',
                                                   remedy_version_response)
                if (len(remedy_version_number) > 0):
                    remedy_version_details.append(remedy_version_number.pop(0))
                else:
                    remedy_version_details.append("-")
            return remedy_version_details;


        # ------------------------------------------------- MAIN CODE -----------------------------------------

        #  Open and Read Component and Version ID Files
        componentID = open("component_ID.txt").readlines()
        component_version_ID = open("component_version_ID.txt").readlines()
        componentName = open("component_name.txt").readlines()

        counter = 0
        risk_profile = []
        remedy_version_no = []
        while counter < (len(componentID)):
            # Risk Profile
            risk_profile_url = 'https://demo.yourblackducksite.com/api/components/' + (
                componentID[counter].rstrip()) + '/versions/' + (component_version_ID[counter].rstrip()) + '/risk-profile'
            risk_profile_url_response = requests.get(risk_profile_url, headers=headers_2)
            # print (risk_profile_url_response.text)

            # Redediate Version
            remedy_versoin_url = 'https://demo.yourblackducksite.com/api/components/' + (
                componentID[counter].rstrip()) + '/versions/' + (component_version_ID[counter].rstrip()) + '/remediating'
            remedy_version_url_response = requests.get(remedy_versoin_url, headers=headers_2)

            # Calling Function to parse risk info
            risk_profile_details = []
            risk_profile_details = compoent_risk_profile((risk_profile_url_response.text))
            risk_profile_details.insert(0, (componentName[counter].rstrip()))
            risk_profile.append(risk_profile_details)

            # Calling Function to parse Remediation version
            remedy_version_details = []
            remedy_version_details = remedy_version((remedy_version_url_response.text))
            remedy_version_no.append(remedy_version_details.pop(0))
            print(componentName[counter].rstrip())
            # print (risk_profile_details)
            # print (remedy_version_details)
            counter += 1

        # ------------------------------------------------------ End of Phase 2 -------------------------------------------
        from openpyxl import Workbook

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Direct Dependencies"
        ws2 = wb.create_sheet ("Transitive Dependencies")

        ws1.append(
            ['APPLICATION NAME', 'APPLICATION BUILD VERSION', 'COMPONENT NAME', 'COMPONENT INSTALLED VERSION', 'COMPONENT OPERATIONAL RISK', 'VULN HIGH', 'VULN MEDIUM', 'VULN LOW', 'CVSS RISK SCORE', 'COMPONENT RELEASE DATE', 'NEW REVISIONS OF COMPONENT AVAILABLE', 'COMPONENT LAST UPDATE DATE', 'COMPONENT AGE IN YEARS', 'BLACKDUCK SCAN TIMESTAMP'])
        ws2.append(
            ['APPLICATION NAME', 'APPLICATION BUILD VERSION', 'COMPONENT NAME', 'COMPONENT INSTALLED VERSION', 'COMPONENT OPERATIONAL RISK', 'VULN HIGH', 'VULN MEDIUM', 'VULN LOW', 'CVSS RISK SCORE', 'COMPONENT RELEASE DATE', 'NEW REVISIONS OF COMPONENT AVAILABLE', 'COMPONENT LAST UPDATE DATE', 'COMPONENT AGE IN YEARS', 'BLACKDUCK SCAN TIMESTAMP'])
        

        counterr = 0
        operational_risk_profile = ()
        vulnerability_risk_profile = ()
        license_risk_profile = ()

        while counterr < (len(bom_components)):
            operational_risk_profile = (
                re.findall(r'"HIGH":(.*?),"MEDIUM":(.*?),"LOW":(.*?),', bom_risk_operational_profile[counterr][3], re.M))
            vulnerability_risk_profile = (
                re.findall(r'"HIGH":(.*?),"MEDIUM":(.*?),"LOW":(.*?),', bom_risk_vuln_profile[counterr][3], re.M))
            license_risk_profile = (
                re.findall(r'"HIGH":(.*?),"MEDIUM":(.*?),"LOW":(.*?),', bom_risk_license_profile[counterr][5], re.M))
            vul_score = (10 * int(vulnerability_risk_profile[0][0]) + 7 * int(vulnerability_risk_profile[0][1]) + 5 * int(
                vulnerability_risk_profile[0][2]))

            from datetime import date

            today_date = date.today()

            from datetime import datetime

            release_date = datetime.strptime(risk_profile[counterr][1], "%Y-%m-%d")
            # print(release_date)
            todays_date = datetime.strptime(str(today_date), "%Y-%m-%d")
            # print(todays_date)
            days_diff = abs((todays_date - release_date).days)
            years_old = int(days_diff / 365)

            # GET CVSS SCORE
            temp_str = bom_components[counterr][3] + '-' + bom_components[counterr][7]

            for cname in cvss_comp_name:
                if cname == temp_str:
                    comp_cvss_base_score_index = cvss_comp_name.index(temp_str)
                    comp_cvss_base_score = cvss_high_score[comp_cvss_base_score_index][1]
                    break
                else:
                    comp_cvss_base_score = '0'
            
            ops_risk = ""
            if (int(operational_risk_profile[0][0]) == 1):
                ops_risk = "HIGH"
            elif (int(operational_risk_profile[0][1]) == 1):
                ops_risk = "MEDIUM"
            elif (int(operational_risk_profile[0][2]) == 1):
                ops_risk = "LOW"
            else:
                ops_risk = "NONE"        
            

            if (str(bom_components[counterr][11])== "FILE_DEPENDENCY_TRANSITIVE"):
                ws2.append(
                    [str(project_names[int(project_index_number)][0]), str(version_names[0][1]), str(bom_components[counterr][3]),
                     (bom_components[counterr][7]), ops_risk, int(vulnerability_risk_profile[0][0]),
                     int(vulnerability_risk_profile[0][1]), int(vulnerability_risk_profile[0][2]), comp_cvss_base_score,
                     str(risk_profile[counterr][1]), int(risk_profile[counterr][2]), str(risk_profile[counterr][4]),
                     years_old, str(version_names[0][0])
                     ])
            else:
                ws1.append(
                    [str(project_names[int(project_index_number)][0]), str(version_names[0][1]), str(bom_components[counterr][3]),
                     (bom_components[counterr][7]), ops_risk, int(vulnerability_risk_profile[0][0]),
                     int(vulnerability_risk_profile[0][1]), int(vulnerability_risk_profile[0][2]), comp_cvss_base_score,
                     str(risk_profile[counterr][1]), int(risk_profile[counterr][2]), str(risk_profile[counterr][4]),
                     years_old, str(version_names[0][0])
                     ])       


            # ws1.append([str(risk_profile[counterr][0])])
            counterr += 1

        # File Operations
        os.remove("bom.txt")
        os.remove("component_ID.txt")
        os.remove("component_version_ID.txt")
        os.remove("component_name.txt")
        os.remove("vulbom.txt")

        # Excel Operation
        excel_name = str(project_names[int(project_index_number)][0]) + '_Blackduck_Security_Risk_Report.xlsx'
        wb.save(excel_name)

        # Sorting Operations
        book = load_workbook(excel_name)
        writer = pd.ExcelWriter(excel_name, engine='openpyxl')
        direct_depen_sheet = pd.read_excel(excel_name, sheet_name="Direct Dependencies")
        transitive_depen_sheet = pd.read_excel(excel_name, sheet_name="Transitive Dependencies")

        sorted_direct = direct_depen_sheet.sort_values(by=['CVSS RISK SCORE', 'COMPONENT OPERATIONAL RISK', 'COMPONENT AGE IN YEARS'], ascending=False, ignore_index=True)
        sorted_transitive = transitive_depen_sheet.sort_values(['CVSS RISK SCORE', 'COMPONENT OPERATIONAL RISK', 'COMPONENT AGE IN YEARS'], ascending=False, ignore_index=True)

        sorted_direct.to_excel(writer, sheet_name='Direct Dependencies')
        sorted_transitive.to_excel(writer, sheet_name='Transitive Dependencies')

        writer.save()
        
        
        #Formatting Operation
        writer = pd.ExcelWriter(excel_name, engine='xlsxwriter')
        direct_depen_sheet = pd.read_excel(excel_name, sheet_name="Direct Dependencies")
        transitive_depen_sheet = pd.read_excel(excel_name, sheet_name="Transitive Dependencies")

        #3
        direct_depen_sheet.to_excel(writer, sheet_name="Direct Dependencies")
        transitive_depen_sheet.to_excel(writer, sheet_name="Transitive Dependencies")

        #4
        workbook  = writer.book
        worksheet1 = writer.sheets['Direct Dependencies']
        worksheet2 = writer.sheets['Transitive Dependencies']

        #5
        format1 = workbook.add_format()
        format1.set_align('left')
        format1.set_align('vcenter')

        format2 = workbook.add_format()
        format2.set_align('center')
        format2.set_align('vcenter')

        format3 = workbook.add_format()
        format3.set_text_wrap()

        format4 = workbook.add_format()
        format4.set_bold()
        format4.set_font_color('red')

        #6
        worksheet1.set_row(0, None, format3)
        worksheet1.set_column('K:K', None, format4)
        worksheet1.set_column('A:B', None, format1, {'hidden': 1})
        worksheet1.set_column('B:E', None, format1)
        worksheet1.set_column('F:O', None, format2)

        worksheet2.set_row(0, None, format3)
        worksheet1.set_column('K:K', None, format4)
        worksheet2.set_column('A:B', None, format1, {'hidden': 1})
        worksheet2.set_column('B:E', None, format1)
        worksheet2.set_column('F:O', None, format2)

        #7
        writer.save()

        print("Blackduck Analysis Completed. Please check the result sheet!")

        project_index_number = project_index_number + 1
    exit()

else:
    print (bdresponse.text)
    print("Error! Check your API Key and please try again later..")
    exit()